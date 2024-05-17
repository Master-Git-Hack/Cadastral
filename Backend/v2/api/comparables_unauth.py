from base64 import b64encode
from datetime import datetime, timedelta
from os import remove

from dateparser import parse
from fastapi import APIRouter, Depends, Request
from num2words import num2words
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Font, PatternFill, Side
from requests import get
from sqlalchemy.orm import Session
from sqlalchemy.sql import func

from .. import config, database, logger
from ..controllers.comparables_catcom import ComparablesCatComReport
from ..middlewares import Middlewares as __Middlewares
from ..middlewares.auth import required
from ..models.cedula_comparables import CedulaComparables
from ..models.cedula_mercado import CedulaMercado

# from ..middlewares.auth import required
from ..models.comparables_catcom import ComparablesCatCom
from ..utils.local import as_complete_date

__response = __Middlewares.Responses()
ua_comparables = APIRouter(
    prefix="/un-auth/comparables",
    tags=["Comparables Catastrales/Comerciales"],
    dependencies=[],
    responses={404: {"description": "Not found"}},
)


@ua_comparables.get("/{username}/cedulas", deprecated=True)
async def get_cedulas(
    username: str,
    db: Session = Depends(database.valuaciones),
):
    """
    Get cedulas
    """
    cedula = CedulaMercado(db)
    if cedula.filter_group(usuario=username) is None:
        return __response.error(message="No se encontraron cedulas", status_code=404)

    return __response.success(
        data=[
            cedula.to_dict(data)
            for data in cedula.current
            if data.fecha + timedelta(days=180) > datetime.now()
        ]
    )


@ua_comparables.get("/{username}/cedula/{id}", deprecated=True)
async def get_cedula_by_id(
    id: int,
    username: str,
    db: Session = Depends(database.valuaciones),
):
    """
    Get cedula by id
    """

    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=username) is None:
        return __response.error(message="No se encontraron cedulas", status_code=404)
    return __response.success(data=cedula.to_dict())


@ua_comparables.post("/{username}/cedula/{id}/reporte/{as_report}", deprecated=True)
async def get_cedula_reporte_by_id(
    id: int,
    username: str,
    db: Session = Depends(database.valuaciones),
):
    """
    Get cedula by id
    """

    # data = await request.json()
    # print(data)
    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=username) is None:
        return __response.error(message="No se encontraron cedulas", status_code=404)
    print(cedula.current.__dict__)
    return __response.success(data=cedula.to_dict())


@ua_comparables.post("/{username}/cedula/{registro}", deprecated=True)
async def create_cedula(
    registro: str,
    username: str,
    db: Session = Depends(database.valuaciones),
):
    """
    Create cedula
    """

    cedula = CedulaMercado(db)
    if cedula.create(registro=registro, usuario=username) is None:
        return __response.error(message="No se pudo crear la cedula", status_code=404)
    return __response.success(data=cedula.to_dict())


@ua_comparables.patch("/{username}/cedula/{id}", deprecated=True)
async def update_cedula(
    id: int,
    username: str,
    request: Request,
    db: Session = Depends(database.valuaciones),
):
    """
    Update cedula
    """
    data = await request.json()

    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=username) is None:
        return __response.error(
            message="No se pudo encontrar la cedula", status_code=404
        )
    if cedula.update(**data) is None:
        return __response.error(
            message="No se pudo actualizar la cedula", status_code=404
        )
    return __response.success(data=cedula.to_dict())


@ua_comparables.delete("/{username}/cedula/{id}", deprecated=True)
async def delete_cedula(
    id: int,
    username: str,
    db: Session = Depends(database.valuaciones),
):
    """
    Delete cedula
    """

    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=username) is None:
        return __response.error(
            message="No se pudo encontrar la cedula", status_code=404
        )
    if cedula.delete() is None:
        return __response.error(
            message="No se pudo eliminar la cedula", status_code=404
        )
    return __response.success(data=cedula.to_dict())


@ua_comparables.get("/{cedula_mercado}", deprecated=True)
async def get_comparables(
    cedula_mercado: int,
    db: Session = Depends(database.valuaciones),
):
    """
    Get comparables
    """

    comp = CedulaComparables(db)
    if comp.filter_group(id_cedula_mercado=cedula_mercado) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    return __response.success(data=comp.to_list())


@ua_comparables.get("/comparable/{id}", deprecated=True)
async def get_comparable_by_id(
    id: int,
    db: Session = Depends(database.valuaciones),
):
    """
    Get comparable by id
    """

    comp = CedulaComparables(db)
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    return __response.success(data=comp.to_dict())


@ua_comparables.post(
    "/comparable/{cedula_mercado}/{tipo}/{comparable}", deprecated=True
)
async def create_comparable(
    cedula_mercado: int,
    tipo: str,
    comparable: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
):
    """
    Create comparable
    """

    data = await request.json()
    comp = CedulaComparables(db)
    if (
        comp.filter(
            id_cedula_mercado=cedula_mercado, id_comparable_catcom=comparable, tipo=tipo
        )
        is not None
    ):
        return __response.error(message="Ya existe el comparable", status_code=404)
    if (
        comp.create(
            tipo=tipo,
            id_cedula_mercado=cedula_mercado,
            id_comparable_catcom=comparable,
            **data,
        )
        is None
    ):
        return __response.error(message="No se pudo crear la cedula", status_code=404)
    return __response.success(data=comp.to_dict())


@ua_comparables.patch("/comparable/{id}", deprecated=True)
async def update_comparable(
    id: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
):
    """
    Update comparable
    """
    data = await request.json()

    comp = CedulaComparables(db)
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if comp.update(**data) is None:
        return __response.error(
            message="No se pudo actualizar la cedula", status_code=404
        )
    return __response.success(data=comp.to_dict())


@ua_comparables.delete("/comparable/{id}", deprecated=True)
async def delete_comparable(
    id: int,
    db: Session = Depends(database.valuaciones),
):
    """
    Delete comparable
    """

    comp = CedulaComparables(db)
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if comp.delete(id) is None:
        return __response.error(
            message="No se pudo eliminar la cedula", status_code=404
        )
    return __response.success(data=comp.to_dict())


@ua_comparables.post("/xlsx/{cedula_mercado}", deprecated=True)
async def generate_xlsx(
    cedula_mercado: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
):

    comp = ComparablesCatCom(db)
    cedula = CedulaComparables(db)
    mercado = CedulaMercado(db)
    if mercado.get(cedula_mercado) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if len(cedula.filter_group(id_cedula_mercado=cedula_mercado)) == 0:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )

    registro = mercado.current.registro
    mercado = mercado.to_dict(exclude=["id", "usuario", "fecha"])
    cedulas = [
        {**comp.to_dict(), "tipo": c.tipo, **mercado}
        for c in cedula.current
        if comp.get(c.id_comparable_catcom) is not None
    ]
    try:
        data = await request.json()

        if not "ids" in data:
            raise Exception("No se encontraron comparables")
        data = [c for c in cedulas if c.get("id") in data["ids"]]
    except Exception as e:
        pass
    data = [
        {"tipo": tipo, "records": [c for c in cedulas if c.get("tipo") == tipo]}
        for tipo in set(c.get("tipo") for c in cedulas)
    ]
    # Crear un nuevo libro de trabajo de Excel
    workbook = Workbook()
    mercado_sheet = workbook.active
    mercado_sheet.title = "MERCADO"
    # cedula_sheet = workbook.create_sheet("CÉDULAS DE MERCADO")
    # Definir estilos
    text = dict(
        header=dict(
            black=Font(name="Arial", size=18, bold=True),
            white=Font(name="Arial", size=18, bold=True, color="FFFFFF"),
        ),
        title=dict(
            black=Font(name="Arial", size=14, bold=True),
            white=Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        ),
        sub_title=dict(
            black=Font(name="Arial", size=12, bold=True),
            white=Font(name="Arial", size=12, bold=True, color="FFFFFF"),
        ),
        normal=dict(
            black=Font(name="Arial", size=9),
            white=Font(name="Arial", size=9, color="FFFFFF"),
        ),
        bold=dict(
            black=Font(name="Arial", size=9, bold=True),
            white=Font(name="Arial", size=9, bold=True, color="FFFFFF"),
        ),
    )
    background = dict(
        red=PatternFill(start_color="F87171", end_color="F87171", fill_type="solid"),
        green=PatternFill(start_color="4ADE80", end_color="4ADE80", fill_type="solid"),
        violet=PatternFill(start_color="A78BFA", end_color="A78BFA", fill_type="solid"),
        kaki=PatternFill(start_color="F3F37A", end_color="F3F37A", fill_type="solid"),
        teal=PatternFill(start_color="2DD4D4", end_color="2DD4D4", fill_type="solid"),
        blue=PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid"),
        pink=PatternFill(start_color="F472B6", end_color="F472B6", fill_type="solid"),
        emerald=PatternFill(
            start_color="A7F3D0", end_color="A7F3D0", fill_type="solid"
        ),
        yellow=PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid"),
    )
    border = dict(
        top=Border(top=Side(border_style="thin", color="000000")),
        bottom=Border(bottom=Side(border_style="thin", color="000000")),
        left=Border(left=Side(border_style="thin", color="000000")),
        right=Border(right=Side(border_style="thin", color="000000")),
        x=Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
        ),
        y=Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        ),
        full=Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        ),
    )

    url_base = "http://172.31.113.151/comparables/imagenes"

    for index, d in enumerate(data):  # replace test for data
        cell = None
        mercado_sheet.append([d.get("tipo")])
        mercado_sheet.merge_cells(
            start_row=mercado_sheet.max_row,
            start_column=1,
            end_row=mercado_sheet.max_row,
            end_column=53,
        )  # Fusionar celdas
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_sheet.max_row, column=1).coordinate
        ]
        cell.font = text["header"]["white"]  # Establecer negrita
        if d.get("tipo") == "VENTA":
            cell.fill = background["green"]  # Establecer color de fondo
        elif d.get("tipo") == "RENTA":
            cell.fill = background["violet"]  # Establecer color de fondo
        else:
            cell.fill = background["red"]  # Establecer color de fondo
        mercado_sheet.append(
            [
                "Datos de Verificación",
                None,
                None,
                None,
                None,
                None,
                None,
                "Ubicación",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Características de Terreno",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Características de Construcción",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Infraestructura",
                None,
                "Valores",
                None,
                None,
                None,
                None,
                None,
                None,
                "Vigencia",
                None,
                None,
                "Elaboró",
            ]
        )
        mercado_last_row = mercado_sheet.max_row
        # Combinar celdas según las especificaciones dadas en el código JSX
        mercado_sheet.merge_cells(f"A{mercado_last_row}:G{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=1).coordinate
        ]
        cell.font = text["sub_title"]["black"]  # Establecer negrita
        cell.fill = background["kaki"]  # Establecer color de fondo

        # Datos de Verificación
        mercado_sheet.merge_cells(f"H{mercado_last_row}:Q{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=8).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["violet"]
        # Ubicación

        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=18).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        # Celda vacía
        mercado_sheet.merge_cells(f"S{mercado_last_row}:AD{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=19).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["red"]
        # Características de Terreno
        mercado_sheet.merge_cells(f"AE{mercado_last_row}:AN{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=31).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["teal"]
        # Características de Construcción
        mercado_sheet.merge_cells(f"AO{mercado_last_row}:AP{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=41).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["kaki"]
        # Infraestructura
        mercado_sheet.merge_cells(f"AQ{mercado_last_row}:AV{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=43).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["blue"]
        # VALORES
        mercado_sheet.merge_cells(
            f"AW{mercado_last_row}:AZ{mercado_last_row}"
        )  # Características de Construcción
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=49).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["pink"]
        mercado_sheet.append(
            [
                "Folio",
                "Tipo de Inmueble",
                "Tipo de Operación",
                "Fecha de Captura",
                "URL Fuente",
                "Informante",
                "Teléfono de Informante",
                "Coordenada UTM X",
                "Coordenada UTM Y",
                "Estado",
                "Municipio",
                "Ciudad o Población",
                "Tipo y Nombre de Colonia / Asentamiento",
                "Tipo y Nombre de la Calle",
                "No. Exterior",
                "No. Interior",
                "Nombre Edificio / Proto / Predio",
                "Régimen de Propiedad",
                "Clasificación Periférica",
                "Clasificación Económica de la Zona (Campo)",
                "Uso de Suelo Carta, Uso Plan",
                "Entre Calles",
                "Ubicación en la Manzana",
                "Número de Frentes",
                "Superficie Terreno M2",
                "Frente ML",
                "Frente Tipo ML",
                "Fondo",
                "Forma",
                "Topografía",
                "Superficie Construcción M2",
                "Proyecto",
                "Edo. Conservación",
                "Tipo de Construcción",
                "Calidad",
                "Edad",
                "Niveles",
                "Unidades Rentables",
                "Descripción de Espacios",
                "T / C",
                "Servicios",
                "Descripción de Servicios",
                "Precio",
                "Precio Unitario",
                "Precio Total USD",
                "Precio Unitario USD",
                "Precio Total Aplicable en la Homologación MXN",
                "Precio Unitario Aplicable en la Homologación MXN",
                "Observaciones",
                "Hoy",
                "Días",
                "Caduca en 6 meses Fecha",
                "Elaboró",
            ]
        )
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_sheet.max_row, column=1).coordinate
        ]
        cell.font = text["bold"]["black"]
        cell.border = border["full"]
        for r in d["records"]:
            if r.get("imagen_1") is not None:
                r["imagen_1"] = f"{url_base}/{r['imagen_1']}"
            else:
                r["imagen_1"] = ""
            if r.get("imagen_2") is not None:
                r["imagen_2"] = f"{url_base}/{r['imagen_2']}"
            else:
                r["imagen_2"] = ""
            if r.get("captura_pantalla") is not None:
                r["captura_pantalla"] = f"{url_base}/{r['captura_pantalla']}"
            else:
                r["captura_pantalla"] = ""

            r["fecha_captura"] = as_complete_date(r.get("fecha_captura", "hoy"))

            fecha_actual = datetime.now()

            # Formatear la fecha actual
            fecha_actual_formato = fecha_actual.strftime("%d de %m del %Y")

            # Obtener la fecha de captura (asumiendo que 'fecha_captura' es una cadena de texto)
            fecha_captura = parse(r.get("fecha_captura", "hoy"))
            # Handle the case where fecha_captura is None
            fecha_captura_formato = ""

            fecha_seis_meses_mas_formato = ""
            mercado_sheet.append(
                [
                    r.get("id"),
                    r.get("tipo_inmueble"),
                    r.get("tipo_operacion"),
                    r.get("fecha_captura"),
                    r.get("url_fuente"),
                    r.get("nombre_anunciante"),
                    r.get("telefono_anunciante"),
                    r.get("x_utm"),
                    r.get("y_utm"),
                    r.get("estado"),
                    r.get("municipio"),
                    r.get("localidad"),
                    f"{r.get('tipo_asentamiento')} {r.get('nombre_asentamiento')}",
                    f"{r.get('tipo_vialidad')} {r.get('nombre_vialidad')}",
                    r.get("numero_exterior"),
                    r.get("numero_interior"),
                    r.get("edificio"),
                    r.get("regimen_propiedad"),
                    r.get("tipo_zona"),
                    r.get("regimen_propiedad"),
                    r.get("tipo_zona"),
                    r.get("uso_suelo_observado"),
                    r.get("uso_suelo_oficial"),
                    r.get("entrecalles"),
                    r.get("ubicacion_manzana"),
                    f"{r.get('numero_frentes')}",  # ({num2words(r.get('numero_frentes',"0"), lang='es').upper()})",
                    r.get("superficie_terreno"),
                    r.get("longitud_frente"),
                    r.get("longitud_frente_tipo"),
                    r.get("longitud_fondo"),
                    r.get("forma"),
                    r.get("topografia"),
                    r.get("superficie_construccion"),
                    r.get("calidad_proyecto"),
                    r.get("estado_conservacion"),
                    r.get("tipo_construccion"),
                    r.get("calidad_construccion"),
                    r.get("edad"),
                    r.get("niveles"),
                    r.get("unidades_rentables"),
                    r.get("descripcion_espacios"),
                    r.get("superficie_terreno", 0)
                    / (r.get("superficie_construccion", 1.0) or 1),
                    r.get(""),
                    r.get("descripcion_espacios"),
                    r.get("valor_total_mercado"),
                    (
                        (
                            r.get("valor_renta", 0)
                            if r.get("tipo") == "RENTA"
                            else r.get("valor_total_mercado", 0) or 0
                        )
                    )
                    / (r.get("superficie_terreno", 1) or 1),
                    r.get("precio_dolar", "-"),
                    (
                        r.get("precio_dolar") / r.get("superficie_terreno", 1)
                        if r.get("precio_dolar")
                        else "-"
                    ),
                    # "-",
                    # "-",
                    r.get("observaciones"),
                    fecha_actual_formato,
                    r.get("fecha_captura"),
                    fecha_seis_meses_mas_formato,
                    r.get("usuario"),
                ]
            )
            for i in range(1, 53):
                current_cell = mercado_sheet.cell(row=mercado_sheet.max_row, column=i)
                current_cell.border = border["full"]
                current_cell.font = text["normal"]["black"]
                current_cell.fill = background["emerald"]

        mercado_sheet.append([])  # Espacio

    mercado_width = {
        "A": 92,
        "B": 183,
        "C": 90,
        "D": 186,
        "E": 90,
        "F": 237,
        "G": 94,
        "H": 92,
        "I": 92,
        "J": 108,
        "K": 209,
        "L": 90,
        "M": 208,
        "N": 136,
        "O": 90,
        "P": 90,
        "Q": 103,
        "R": 136,
        "S": 94,
        "T": 239,
        "U": 215,
        "V": 204,
        "W": 150,
        "X": 76,
        "Y": 101,
        "Z": 76,
        "AA": 76,
        "AB": 76,
        "AC": 125,
        "AD": 162,
        "AE": 105,
        "AF": 90,
        "AG": 149,
        "AH": 149,
        "AI": 104,
        "AJ": 90,
        "AK": 90,
        "AL": 90,
        "AM": 183,
        "AN": 90,
        "AO": 143,
        "AP": 543,
        "AQ": 120,
        "AR": 120,
        "AS": 120,
        "AT": 120,
        "AU": 120,
        "AV": 120,
        "AW": 246,
        "AX": 79,
        "AY": 72,
        "AZ": 183,
        "BA": 77,
    }

    for column, width in mercado_width.items():
        mercado_sheet.column_dimensions[column].width = width / 7.5
    # for column, width in cedula_width.items():
    #     cedula_sheet.column_dimensions[column].width = width / 7.5
    filename = f"{registro}.xlsx"
    path = f"{config.PATHS.tmp}/{filename}"
    workbook.save(path)
    # filename, path = create_file(data)
    return __response.send_file(filename=filename, path=path, delete=True)


from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from PIL import Image


@ua_comparables.get("/preview/image/{comparable}/{width}/{height}", deprecated=True)
async def handle_image_preview(
    comparable: int,
    width: int = 200,
    height: int = 200,
    db: Session = Depends(database.valuaciones),
    # ,
):
    comp = ComparablesCatCom(db)
    if comp.get(comparable) is None:
        return __response.error(message="Registro No encontrado")
    filename = comp.current.imagen_2
    if filename is None:
        return __response.error(message="Imagen no encontrada")
    url_base = "http://172.31.113.151/comparables/imagenes"
    ext = filename.split(".")[-1]
    image_url = f"{url_base}/{filename}"

    try:
        response = get(image_url)
        # response.raise_for_status()
        path = f"{config.PATHS.tmp}/{filename}"

        with open(path, "wb") as f:
            f.write(response.content)
        img = Image.open(path)
        img_width, img_height = img.size
        left = (img_width - width) // 2
        top = (img_height - height) // 2
        right = (img_width + width) // 2
        bottom = (img_height + height) // 2
        cropped = img.crop((left, top, right, bottom))
        path = path.replace(ext, f"_crop.{ext}")
        cropped.resize((img_width, img_height), Image.Resampling.LANCZOS)
        cropped.save(path, quality=95)
        return __response.send_file(
            filename=filename, path=path, media_type=f"image/{ext}", delete=True
        )
    except Exception as e:
        return __response.error(
            message="No se pudo descargar la imagen", status_code=404
        )


@ua_comparables.get("/image/{filename}/{crop}", deprecated=True)
async def handle_images(
    filename: str,
    crop: bool = False,
    # ,
):
    # if isinstance(user, dict):
    #     return __response.error(**user)
    if filename == "null" or filename is None:
        return __response.error("No se encontró la imagen", status_code=404)
    # download image an save on config.paths.tmp and send as a response.send_file
    url_base = "http://172.31.113.151/comparables/imagenes"
    ext = filename.split(".")[-1]
    image_url = f"{url_base}/{filename}"

    try:
        response = get(image_url)
        # response.raise_for_status()
        path = f"{config.PATHS.tmp}/{filename}"

        with open(path, "wb") as f:
            f.write(response.content)
        if crop:
            img = Image.open(path)
            width, height = img.size
            left = (width - 200) // 2
            top = (height - 200) // 2
            right = (width + 200) // 2
            bottom = (height + 200) // 2
            cropped = img.crop((left, top, right, bottom))
            path = path.replace(ext, f"_crop.{ext}")
            cropped.resize((width, height), Image.Resampling.LANCZOS)
            cropped.save(path, quality=95)

        return __response.send_file(
            filename=filename, path=path, media_type=f"image/{ext}", delete=True
        )
    except Exception as e:
        return __response.error(
            message="No se pudo descargar la imagen", status_code=404
        )


@ua_comparables.post("/preview/{cedula_mercado}", deprecated=True)
async def generate_preview(
    cedula_mercado: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
):

    comp = ComparablesCatCom(db)
    cedula = CedulaComparables(db)
    mercado = CedulaMercado(db)
    if mercado.get(cedula_mercado) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if len(cedula.filter_group(id_cedula_mercado=cedula_mercado)) == 0:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    mercado = mercado.to_dict(exclude=["id", "usuario", "fecha"])
    cedulas = [
        {**comp.to_dict(), "tipo": c.tipo, **mercado}
        for c in cedula.current
        if comp.get(c.id_comparable_catcom) is not None
    ]
    try:
        data = await request.json()

        if not "ids" in data:
            raise Exception("No se encontraron comparables")
        data = [c for c in cedulas if c.get("id") in data["ids"]]
    except Exception as e:
        pass
    data = [
        {"tipo": tipo, "records": [c for c in cedulas if c.get("tipo") == tipo]}
        for tipo in set(c.get("tipo") for c in cedulas)
    ]
    # url_base = "http://172.31.113.151/comparables/imagenes"
    for d in data:
        for r in d["records"]:
            # r["imagen_1"] = f"{url_base}/{r['imagen_1']}"
            # r["imagen_2"] = f"{url_base}/{r['imagen_2']}"
            # r["captura_pantalla"] = f"{url_base}/{r['captura_pantalla']}"
            r["fecha_captura"] = as_complete_date(r["fecha_captura"])

    return __response.success(data=data)
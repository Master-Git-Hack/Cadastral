from datetime import datetime, timedelta

from dateparser import parse
from fastapi import APIRouter, Depends, Request
from num2words import num2words
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Font, PatternFill, Side
from requests import get
from sqlalchemy.orm import Session
from sqlalchemy.sql import func

from .. import config, database, logger
from ..controllers.comparables_catcom import ComparablesCatComReport
from ..middlewares import Middlewares as __Middlewares
from ..middlewares.auth import required
from ..models.cedula_comparables import CedulaComparables
from ..models.cedula_mercado import CedulaMercado
# from ..middlewares.auth import required
from ..models.comparables_catcom import ComparablesCatCom
from ..utils.local import as_complete_date

__response = __Middlewares.Responses()
comparables = APIRouter(
    prefix="/comparables",
    tags=["Comparables Catastrales/Comerciales"],
    dependencies=[],
    responses={404: {"description": "Not found"}},
)


@comparables.get("/cedulas")
async def get_cedulas(
    db: Session = Depends(database.valuaciones), user=Depends(required)
):
    """
    Get cedulas
    """
    if isinstance(user, dict):
        return __response.error(**user)
    cedula = CedulaMercado(db)
    if cedula.filter_group(usuario=user.usuario) is None:
        return __response.error(message="No se encontraron cedulas", status_code=404)

    return __response.success(
        data=[
            cedula.to_dict(data)
            for data in cedula.current
            if data.fecha + timedelta(days=180) > datetime.now()
        ]
    )


@comparables.get("/cedula/{id}")
async def get_cedula_by_id(
    id: int, db: Session = Depends(database.valuaciones), user=Depends(required)
):
    """
    Get cedula by id
    """
    if isinstance(user, dict):
        return __response.error(**user)
    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=user.usuario) is None:
        return __response.error(message="No se encontraron cedulas", status_code=404)
    return __response.success(data=cedula.to_dict())


@comparables.post("/cedula/{id}/reporte/{as_report}")
async def get_cedula_reporte_by_id(
    id: int,
    request: Request,
    as_report: str = "mercado",
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Get cedula by id
    """
    if isinstance(user, dict):
        return __response.error(**user)

    # data = await request.json()
    # print(data)
    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=user.usuario) is None:
        return __response.error(message="No se encontraron cedulas", status_code=404)
    print(cedula.current.__dict__)
    return __response.success(data=cedula.to_dict())


@comparables.post("/cedula/{registro}")
async def create_cedula(
    registro: str,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Create cedula
    """
    if isinstance(user, dict):
        return __response.error(**user)
    cedula = CedulaMercado(db)
    if cedula.create(registro=registro, usuario=user.usuario) is None:
        return __response.error(message="No se pudo crear la cedula", status_code=404)
    return __response.success(data=cedula.to_dict())


@comparables.patch("/cedula/{id}")
async def update_cedula(
    id: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Update cedula
    """
    data = await request.json()
    if isinstance(user, dict):
        return __response.error(**user)
    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=user.usuario) is None:
        return __response.error(
            message="No se pudo encontrar la cedula", status_code=404
        )
    if cedula.update(**data) is None:
        return __response.error(
            message="No se pudo actualizar la cedula", status_code=404
        )
    return __response.success(data=cedula.to_dict())


@comparables.delete("/cedula/{id}")
async def delete_cedula(
    id: int,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Delete cedula
    """
    if isinstance(user, dict):
        return __response.error(**user)
    cedula = CedulaMercado(db)
    if cedula.filter(id=id, usuario=user.usuario) is None:
        return __response.error(
            message="No se pudo encontrar la cedula", status_code=404
        )
    if cedula.delete() is None:
        return __response.error(
            message="No se pudo eliminar la cedula", status_code=404
        )
    return __response.success(data=cedula.to_dict())


@comparables.get("/{cedula_mercado}")
async def get_comparables(
    cedula_mercado: int,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Get comparables
    """
    if isinstance(user, dict):
        return __response.error(**user)
    comp = CedulaComparables(db)
    if comp.filter_group(id_cedula_mercado=cedula_mercado) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    return __response.success(data=comp.to_list())


@comparables.get("/comparable/{id}")
async def get_comparable_by_id(
    id: int,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Get comparable by id
    """
    if isinstance(user, dict):
        return __response.error(**user)
    comp = CedulaComparables(db)
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    return __response.success(data=comp.to_dict())


@comparables.post("/comparable/{cedula_mercado}/{tipo}/{comparable}")
async def create_comparable(
    cedula_mercado: int,
    tipo: str,
    comparable: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Create comparable
    """
    if isinstance(user, dict):
        return __response.error(**user)
    data = await request.json()
    comp = CedulaComparables(db)
    if (
        comp.filter(
            id_cedula_mercado=cedula_mercado, id_comparable_catcom=comparable, tipo=tipo
        )
        is not None
    ):
        return __response.error(message="Ya existe el comparable", status_code=404)
    if (
        comp.create(
            tipo=tipo,
            id_cedula_mercado=cedula_mercado,
            id_comparable_catcom=comparable,
            **data,
        )
        is None
    ):
        return __response.error(message="No se pudo crear la cedula", status_code=404)
    return __response.success(data=comp.to_dict())


@comparables.patch("/comparable/{id}")
async def update_comparable(
    id: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Update comparable
    """
    data = await request.json()
    if isinstance(user, dict):
        return __response.error(**user)
    comp = CedulaComparables(db)
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if comp.update(**data) is None:
        return __response.error(
            message="No se pudo actualizar la cedula", status_code=404
        )
    return __response.success(data=comp.to_dict())


@comparables.delete("/comparable/{id}")
async def delete_comparable(
    id: int,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    """
    Delete comparable
    """
    if isinstance(user, dict):
        return __response.error(**user)
    comp = CedulaComparables(db)
    if comp.get(id) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if comp.delete(id) is None:
        return __response.error(
            message="No se pudo eliminar la cedula", status_code=404
        )
    return __response.success(data=comp.to_dict())


@comparables.post("/xlsx/{cedula_mercado}")
async def generate_xlsx(
    cedula_mercado: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    if isinstance(user, dict):
        return __response.error(**user)
    comp = ComparablesCatCom(db)
    cedula = CedulaComparables(db)
    mercado = CedulaMercado(db)
    if mercado.get(cedula_mercado) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if len(cedula.filter_group(id_cedula_mercado=cedula_mercado)) == 0:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )

    registro = mercado.current.registro
    mercado = mercado.to_dict(exclude=["id", "usuario", "fecha"])
    cedulas = [
        {**comp.to_dict(), "tipo": c.tipo, **mercado}
        for c in cedula.current
        if comp.get(c.id_comparable_catcom) is not None
    ]
    try:
        data = await request.json()

        if not "ids" in data:
            raise Exception("No se encontraron comparables")
        data = [c for c in cedulas if c.get("id") in data["ids"]]
    except Exception as e:
        pass
    data = [
        {"tipo": tipo, "records": [c for c in cedulas if c.get("tipo") == tipo]}
        for tipo in set(c.get("tipo") for c in cedulas)
    ]
    # Crear un nuevo libro de trabajo de Excel
    workbook = Workbook()
    mercado_sheet = workbook.active
    mercado_sheet.title = "MERCADO"
    cedula_sheet = workbook.create_sheet("CÉDULAS DE MERCADO")
    # Definir estilos
    text = dict(
        header=dict(
            black=Font(name="Arial", size=18, bold=True),
            white=Font(name="Arial", size=18, bold=True, color="FFFFFF"),
        ),
        title=dict(
            black=Font(name="Arial", size=14, bold=True),
            white=Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        ),
        sub_title=dict(
            black=Font(name="Arial", size=12, bold=True),
            white=Font(name="Arial", size=12, bold=True, color="FFFFFF"),
        ),
        normal=dict(
            black=Font(name="Arial", size=9),
            white=Font(name="Arial", size=9, color="FFFFFF"),
        ),
        bold=dict(
            black=Font(name="Arial", size=9, bold=True),
            white=Font(name="Arial", size=9, bold=True, color="FFFFFF"),
        ),
    )
    background = dict(
        red=PatternFill(start_color="F87171", end_color="F87171", fill_type="solid"),
        green=PatternFill(start_color="4ADE80", end_color="4ADE80", fill_type="solid"),
        violet=PatternFill(start_color="A78BFA", end_color="A78BFA", fill_type="solid"),
        kaki=PatternFill(start_color="F3F37A", end_color="F3F37A", fill_type="solid"),
        teal=PatternFill(start_color="2DD4D4", end_color="2DD4D4", fill_type="solid"),
        blue=PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid"),
        pink=PatternFill(start_color="F472B6", end_color="F472B6", fill_type="solid"),
        emerald=PatternFill(
            start_color="A7F3D0", end_color="A7F3D0", fill_type="solid"
        ),
        yellow=PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid"),
    )
    border = dict(
        top=Border(top=Side(border_style="thin", color="000000")),
        bottom=Border(bottom=Side(border_style="thin", color="000000")),
        left=Border(left=Side(border_style="thin", color="000000")),
        right=Border(right=Side(border_style="thin", color="000000")),
        x=Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
        ),
        y=Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        ),
        full=Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        ),
    )

    url_base = "http://172.31.113.151/comparables/imagenes"

    for index, d in enumerate(data):  # replace test for data
        cell = None
        mercado_sheet.append([d.get("tipo")])
        mercado_sheet.merge_cells(
            start_row=mercado_sheet.max_row,
            start_column=1,
            end_row=mercado_sheet.max_row,
            end_column=53,
        )  # Fusionar celdas
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_sheet.max_row, column=1).coordinate
        ]
        cell.font = text["header"]["white"]  # Establecer negrita
        if d.get("tipo") == "VENTA":
            cell.fill = background["green"]  # Establecer color de fondo
        elif d.get("tipo") == "RENTA":
            cell.fill = background["violet"]  # Establecer color de fondo
        else:
            cell.fill = background["red"]  # Establecer color de fondo
        mercado_sheet.append(
            [
                "Datos de Verificación",
                None,
                None,
                None,
                None,
                None,
                None,
                "Ubicación",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Características de Terreno",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Características de Construcción",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "Infraestructura",
                None,
                "Valores",
                None,
                None,
                None,
                None,
                None,
                None,
                "Vigencia",
                None,
                None,
                "Elaboró",
            ]
        )
        mercado_last_row = mercado_sheet.max_row
        # Combinar celdas según las especificaciones dadas en el código JSX
        mercado_sheet.merge_cells(f"A{mercado_last_row}:G{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=1).coordinate
        ]
        cell.font = text["sub_title"]["black"]  # Establecer negrita
        cell.fill = background["kaki"]  # Establecer color de fondo

        # Datos de Verificación
        mercado_sheet.merge_cells(f"H{mercado_last_row}:Q{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=8).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["violet"]
        # Ubicación

        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=18).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        # Celda vacía
        mercado_sheet.merge_cells(f"S{mercado_last_row}:AD{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=19).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["red"]
        # Características de Terreno
        mercado_sheet.merge_cells(f"AE{mercado_last_row}:AN{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=31).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["teal"]
        # Características de Construcción
        mercado_sheet.merge_cells(f"AO{mercado_last_row}:AP{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=41).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["kaki"]
        # Infraestructura
        mercado_sheet.merge_cells(f"AQ{mercado_last_row}:AV{mercado_last_row}")
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=43).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["blue"]
        # VALORES
        mercado_sheet.merge_cells(
            f"AW{mercado_last_row}:AZ{mercado_last_row}"
        )  # Características de Construcción
        cell = mercado_sheet[
            mercado_sheet.cell(row=mercado_last_row, column=49).coordinate
        ]
        cell.font = text["sub_title"]["black"]
        cell.fill = background["pink"]
        for r in d["records"]:
            if r.get("imagen_1") is not None:
                r["imagen_1"] = f"{url_base}/{r['imagen_1']}"
            else:
                r["imagen_1"] = ""
            if r.get("imagen_2") is not None:
                r["imagen_2"] = f"{url_base}/{r['imagen_2']}"
            else:
                r["imagen_2"] = ""
            if r.get("captura_pantalla") is not None:
                r["captura_pantalla"] = f"{url_base}/{r['captura_pantalla']}"
            else:
                r["captura_pantalla"] = ""
            
            r["fecha_captura"] = as_complete_date(r.get("fecha_captura", "hoy"))

            imagen_1 = get(r["imagen_1"]).content
            imagen_1_path = f"{config.PATHS.tmp}/imagen_1.png"
            with open(imagen_1_path, "wb") as f:
                f.write(imagen_1)
            imagen_2 = get(r["imagen_2"]).content
            imagen_2_path = f"{config.PATHS.tmp}/imagen_2.png"
            with open(imagen_2_path, "wb") as f:
                f.write(imagen_2)
            captura_pantalla = get(r["captura_pantalla"]).content
            captura_pantalla_path = f"{config.PATHS.tmp}/captura_pantalla.png"
            with open(captura_pantalla_path, "wb") as f:
                f.write(captura_pantalla)
            mercado_sheet.append(
                [
                    "Folio",
                    "Tipo de Inmueble",
                    "Tipo de Operación",
                    "Fecha de Captura",
                    "URL Fuente",
                    "Informante",
                    "Teléfono de Informante",
                    "Coordenada UTM X",
                    "Coordenada UTM Y",
                    "Estado",
                    "Municipio",
                    "Ciudad o Población",
                    "Tipo y Nombre de Colonia / Asentamiento",
                    "Tipo y Nombre de la Calle",
                    "No. Exterior",
                    "No. Interior",
                    "Nombre Edificio / Proto / Predio",
                    "Régimen de Propiedad",
                    "Clasificación Periférica",
                    "Clasificación Económica de la Zona (Campo)",
                    "Uso de Suelo Carta, Uso Plan",
                    "Entre Calles",
                    "Ubicación en la Manzana",
                    "Número de Frentes",
                    "Superficie Terreno M2",
                    "Frente ML",
                    "Frente Tipo ML",
                    "Fondo",
                    "Forma",
                    "Topografía",
                    "Superficie Construcción M2",
                    "Proyecto",
                    "Edo. Conservación",
                    "Tipo de Construcción",
                    "Calidad",
                    "Edad",
                    "Niveles",
                    "Unidades Rentables",
                    "Descripción de Espacios",
                    "T / C",
                    "Servicios",
                    "Descripción de Servicios",
                    "Precio",
                    "Precio Unitario",
                    "Precio Total USD",
                    "Precio Unitario USD",
                    "Precio Total Aplicable en la Homologación MXN",
                    "Precio Unitario Aplicable en la Homologación MXN",
                    "Observaciones",
                    "Hoy",
                    "Días",
                    "Caduca en 6 meses Fecha",
                    "Elaboró",
                ]
            )
            cell = mercado_sheet[
                mercado_sheet.cell(row=mercado_sheet.max_row, column=1).coordinate
            ]
            cell.font = text["bold"]["black"]
            cell.border = border["full"]
            # mercado_sheet.append(r["tipo"])
            # mercado_sheet.merge_cells(
            #     start_row=mercado_sheet.max_row,
            #     start_column=1,
            #     end_row=mercado_sheet.max_row,
            #     end_column=3,
            # )  # Fusionar celdas
            # mercado_sheet[mercado_sheet.cell(row=mercado_sheet.max_row, column=1).coordinate].font = (
            #     font_bold  # Establecer negrita
            # )
            # mercado_sheet[mercado_sheet.cell(row=mercado_sheet.max_row, column=1).coordinate].fill = (
            #     fill_gray  # Establecer color de fondo
            # )
            # Obtener la fecha actual
            fecha_actual = datetime.now()

            # Formatear la fecha actual
            fecha_actual_formato = fecha_actual.strftime("%d de %m del %Y")

            # Obtener la fecha de captura (asumiendo que 'fecha_captura' es una cadena de texto)
            fecha_captura = parse(r.get("fecha_captura", "hoy"))
            # Handle the case where fecha_captura is None
            fecha_captura_formato = ""

            fecha_seis_meses_mas_formato = ""
            mercado_sheet.append(
                [
                    r.get("id"),
                    r.get("tipo_inmueble"),
                    r.get("tipo_operacion"),
                    r.get("fecha_captura"),
                    r.get("url_fuente"),
                    r.get("nombre_anunciante"),
                    r.get("telefono_anunciante"),
                    r.get("x_utm"),
                    r.get("y_utm"),
                    r.get("estado"),
                    r.get("municipio"),
                    r.get("localidad"),
                    f"{r.get('tipo_asentamiento')} {r.get('nombre_asentamiento')}",
                    f"{r.get('tipo_vialidad')} {r.get('nombre_vialidad')}",
                    r.get("numero_exterior"),
                    r.get("numero_interior"),
                    r.get("edificio"),
                    r.get("regimen_propiedad"),
                    r.get("tipo_zona"),
                    r.get("regimen_propiedad"),
                    r.get("tipo_zona"),
                    r.get("uso_suelo_observado"),
                    r.get("uso_suelo_oficial"),
                    r.get("entrecalles"),
                    r.get("ubicacion_manzana"),
                    f"{r.get('numero_frentes')} ({num2words(r.get('numero_frentes'), lang='es').upper()})",
                    r.get("superficie_terreno"),
                    r.get("longitud_frente"),
                    r.get("longitud_frente_tipo"),
                    r.get("longitud_fondo"),
                    r.get("forma"),
                    r.get("topografia"),
                    r.get("superficie_construccion"),
                    r.get("calidad_proyecto"),
                    r.get("estado_conservacion"),
                    r.get("tipo_construccion"),
                    r.get("calidad_construccion"),
                    r.get("edad"),
                    r.get("niveles"),
                    r.get("unidades_rentables"),
                    r.get("descripcion_espacios"),
                    r.get("superficie_terreno", 0)
                    / (r.get("superficie_construccion", 1.0) or 1),
                    r.get(""),
                    r.get("descripcion_espacios"),
                    r.get("valor_total_mercado"),
                    (
                        r.get("valor_renta", 0)
                        if r.get("tipo") == "RENTA"
                        else r.get("valor_total_mercado")
                    )
                    / r.get("superficie_terreno", 1),
                    r.get("precio_dolar", "-"),
                    (
                        r.get("precio_dolar") / r.get("superficie_terreno", 1)
                        if r.get("precio_dolar")
                        else "-"
                    ),
                    # "-",
                    # "-",
                    r.get("observaciones"),
                    fecha_actual_formato,
                    r.get("fecha_captura"),
                    fecha_seis_meses_mas_formato,
                    r.get("usuario"),
                ]
            )
            for i in range(1, 53):
                current_cell = mercado_sheet.cell(row=mercado_sheet.max_row, column=i)
                current_cell.border = border["full"]
                current_cell.font = text["normal"]["black"]
                current_cell.fill = background["emerald"]
            # cell = mercado_sheet[
            #     mercado_sheet.cell(row=mercado_sheet.max_row).coordinate
            # ]
            # cell.font = text["normal"]["black"]
            # cell.border = border["full"]
            # cell.fill = background["emerald"]
            cedula_sheet.append([])
            cedula_sheet.append(
                [
                    None,
                    f"COMPARABLES DE {r.get('tipo')}",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:Q{cedula_last_row}")
            cedula_sheet.append([])
            cedula_sheet.append(
                [
                    None,
                    # Image(imagen_1_path, f"B{cedula_last_row}"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    # Image(imagen_2_path, f"J{cedula_last_row}"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            imagen_1 = Image(imagen_1_path)
            imagen_2 = Image(imagen_2_path)
            # imagen_1.anchor = "B{cedula_last_row}"
            # imagen_2.anchor = "J{cedula_last_row}"
            cedula_sheet.add_image(imagen_1)
            cedula_sheet.add_image(imagen_2)

            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:P{cedula_last_row}")

            cedula_sheet.append(
                [
                    None,
                    "COMPARABLE 1",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "MICROLOCALIZACIÓN",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:P{cedula_last_row}")
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Fecha de Captura:",
                    None,
                    r.get("fecha_captura"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Periferia:",
                    None,
                    r.get("tipo_zona"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Tipo de Inmueble:",
                    None,
                    r.get("tipo_inmueble"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Zona Económica:",
                    None,
                    r.get("uso_suelo_observado"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Informante:",
                    None,
                    r.get("nombre_anunciante"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Uso de Suelo:",
                    None,
                    r.get("uso_suelo_oficial"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Telefono del Informante:",
                    None,
                    r.get("telefono_anunciante"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Entre Calles:",
                    None,
                    r.get("entrecalles"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "URL Fuente:",
                    None,
                    r.get("url_fuente"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Ubicación en la MZA:",
                    None,
                    r.get("ubicacion_manzana"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Superficie:",
                    None,
                    r.get("superficie_terreno"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Tipo de Operación:",
                    None,
                    r.get("tipo_operacion"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "No. de Frentes:",
                    None,
                    f'{r.get("numero_frentes")} ({num2words(r.get("numero_frentes"), lang="es").upper()})',
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Frentes ML:",
                    None,
                    r.get("longitud_frente"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW# NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Ubicación",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Frentes Tipo:",
                    None,
                    r.get("longitud_frente_tipo"),
                    None,
                    "Fondo:",
                    r.get("longitud_fondo"),
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:M{cedula_last_row}")
            cedula_sheet.merge_cells(f"O{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Estado:",
                    None,
                    r.get("estado"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Forma:",
                    None,
                    r.get("forma"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Municipio:",
                    None,
                    r.get("municipio"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Topografía:",
                    None,
                    r.get("topografia"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            servicios = "No tiene"
            if (
                r.get("agua")
                or r.get("drenaje")
                or r.get("energia_electrica")
                or r.get("alumbrado_publico")
                or r.get("banqueta")
                or r.get("pavimento")
                or r.get("telefonia")
            ):
                servicios = "Tiene Algunos"

            if (
                r.get("agua")
                and r.get("drenaje")
                and r.get("energia_electrica")
                and r.get("alumbrado_publico")
                and r.get("banqueta")
                and r.get("pavimento")
                and r.get("telefonia")
            ):
                servicios = "Si Tiene Completos"

            cedula_sheet.append(
                [
                    None,
                    "Ciudad/Población:",
                    None,
                    r.get("localidad"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Servicios:",
                    None,
                    servicios,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Asentamiento:",
                    None,
                    f"{r.get('tipo_asentamiento')} {r.get('nombre_asentamiento')}",
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Régimen de Propiedad",
                    None,
                    r.get("regimen_propiedad"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Nombre de la Vialidad:",
                    None,
                    f"{r.get('tipo_vialidad')} {r.get('nombre_vialidad')}",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "No. Exterior:",
                    None,
                    r.get("numero_exterior"),
                    None,
                    "No. Interior:",
                    None,
                    r.get("numero_interior"),
                    None,
                    "Valores",
                    None,
                    r.get("regimen_propiedad"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:E{cedula_last_row}")
            cedula_sheet.merge_cells(f"G{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Edificio Predio Prototipo:",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Precio:",
                    None,
                    r.get("valor_total_mercado"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Coordenadas:",
                    None,
                    "X",
                    r.get("x_utm"),
                    "Y",
                    r.get("y_utm"),
                    None,
                    None,
                    "Precio Unitario:",
                    None,
                    (
                        r.get("valor_renta")
                        if d.get("tipo") == "RENTA"
                        else r.get("valor_total_mercado") / r.get("superficie_terreno")
                    ),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Precio USD:",
                    None,
                    None,
                    "Precio Unitario USD:",
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"M{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "Observaciones:",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    r.get("precio_dolar", "-"),
                    None,
                    "-",
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"M{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    r.get("observaciones"),
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Infraestructura",
                    None,
                    servicios,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:C{cedula_last_row}")
            cedula_sheet.merge_cells(f"D{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:K{cedula_last_row}")
            cedula_sheet.merge_cells(f"L{cedula_last_row}:P{cedula_last_row}")
            # END ROW
            # NEW ROW
            cedula_sheet.append(
                [
                    None,
                    "captura_pantalla",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "Precio:",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            cedula_last_row = cedula_sheet.max_row
            cedula_sheet.merge_cells(f"B{cedula_last_row}:H{cedula_last_row}")
            cedula_sheet.merge_cells(f"J{cedula_last_row}:P{cedula_last_row}")
            # END ROW
        mercado_sheet.append([])  # Espacio
        cedula_sheet.append([])
    # for cell in mercado_sheet[
    #     mercado_sheet.cell(row=mercado_sheet.max_row, column=1)
    #     .coordinate : mercado_sheet.cell(row=mercado_sheet.max_row, column=3)
    #     .coordinate
    # ]:
    #     cell.border = border  # Establecer bordes
    mercado_width = {
        "A": 92,
        "B": 183,
        "C": 90,
        "D": 186,
        "E": 90,
        "F": 237,
        "G": 94,
        "H": 92,
        "I": 92,
        "J": 108,
        "K": 209,
        "L": 90,
        "M": 208,
        "N": 136,
        "O": 90,
        "P": 90,
        "Q": 103,
        "R": 136,
        "S": 94,
        "T": 239,
        "U": 215,
        "V": 204,
        "W": 150,
        "X": 76,
        "Y": 101,
        "Z": 76,
        "AA": 76,
        "AB": 76,
        "AC": 125,
        "AD": 162,
        "AE": 105,
        "AF": 90,
        "AG": 149,
        "AH": 149,
        "AI": 104,
        "AJ": 90,
        "AK": 90,
        "AL": 90,
        "AM": 183,
        "AN": 90,
        "AO": 143,
        "AP": 543,
        "AQ": 120,
        "AR": 120,
        "AS": 120,
        "AT": 120,
        "AU": 120,
        "AV": 120,
        "AW": 246,
        "AX": 79,
        "AY": 72,
        "AZ": 183,
        "BA": 77,
    }
    cedula_width = {
        "A": 13,
        "B": 13,
        "C": 87,
        "D": 100,
        "E": 87,
        "F": 87,
        "G": 87,
        "H": 13,
        "I": 26,
        "J": 86,
        "K": 86,
        "M": 87,
        "N": 87,
        "O": 87,
        "P": 13,
        "Q": 13,
    }
    for column, width in mercado_width.items():
        mercado_sheet.column_dimensions[column].width = width / 7.5
    for column, width in cedula_width.items():
        cedula_sheet.column_dimensions[column].width = width / 7.5
    filename = f"{registro}.xlsx"
    path = f"{config.PATHS.tmp}/{filename}"
    workbook.save(path)
    return __response.send_file(filename=filename, path=path, delete=True)


from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side


@comparables.post("/preview/{cedula_mercado}")
async def generate_preview(
    cedula_mercado: int,
    request: Request,
    db: Session = Depends(database.valuaciones),
    user=Depends(required),
):
    if isinstance(user, dict):
        return __response.error(**user)
    comp = ComparablesCatCom(db)
    cedula = CedulaComparables(db)
    mercado = CedulaMercado(db)
    if mercado.get(cedula_mercado) is None:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    if len(cedula.filter_group(id_cedula_mercado=cedula_mercado)) == 0:
        return __response.error(
            message="No se encontraron comparables", status_code=404
        )
    mercado = mercado.to_dict(exclude=["id", "usuario", "fecha"])
    cedulas = [
        {**comp.to_dict(), "tipo": c.tipo, **mercado}
        for c in cedula.current
        if comp.get(c.id_comparable_catcom) is not None
    ]
    try:
        data = await request.json()

        if not "ids" in data:
            raise Exception("No se encontraron comparables")
        data = [c for c in cedulas if c.get("id") in data["ids"]]
    except Exception as e:
        pass
    data = [
        {"tipo": tipo, "records": [c for c in cedulas if c.get("tipo") == tipo]}
        for tipo in set(c.get("tipo") for c in cedulas)
    ]
    url_base = "http://172.31.113.151/comparables/imagenes"
    for d in data:
        for r in d["records"]:
            r["imagen_1"] = f"{url_base}/{r['imagen_1']}"
            r["imagen_2"] = f"{url_base}/{r['imagen_2']}"
            r["captura_pantalla"] = f"{url_base}/{r['captura_pantalla']}"
            r["fecha_captura"] = as_complete_date(r["fecha_captura"])

    return __response.success(data=data)
    # data = await request.json()
    # # print(data)
    # report = ComparablesCatComReport(db)
    # file = report.preview(
    #     cedula_mercado, as_report=as_report, comparable=comparable, tipo=tipo, **data
    # )

    # if file is None:
    #     return __response.error(
    #         message="Comparable no disponible, ya han transcurrido más de 6 meses.",
    #         status_code=404,
    #     )
    # if not file:
    #     return __response.error(
    #         message="No se pudo generar el reporte", status_code=421
    #     )
    # route = file.split("/")
    # # print("/".join(route[:-1]))
    # # print(route[-1])
    # return __response.send_file(filename=route[-1], path=file, delete=True)


# @comparables.get("/catatastrales_comerciales")
# async def get_comparables_catcom(
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get comparables catrastrales/comerciales
#     """

#     if isinstance(user, dict):
#         return __response.error(**user)
#     comp = ComparablesCatCom(db)
#     if comp.all() is None:
#         return __response.error(
#             message="No se encontraron comparables", status_code=404
#         )
#     return __response.success(data=comp.to_list())


# @comparables.get("/catatastral_comercial/{id}")
# async def get_comparables_catcom_by_id(
#     id: int,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get comparables catrastrales/comerciales by id
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)
#     comp = ComparablesCatCom(db)
#     if comp.get(id) is None:
#         return __response.error(
#             message="No se encontraron comparables", status_code=404
#         )
#     return __response.success(data=comp.to_dict())


# @comparables.post("/catastral_comercial/{cedula_type}/reporte")
# async def get_reporte_catastra_comercial(
#     request: Request,
#     cedula_type: str = "mercado",
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get reporte comparables catrastrales/comerciales
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)
#     # data = await request.json()
#     comp = ComparablesCatComReport(db)
#     # comp.create(type=cedula_type, *data.get("ids", []), **data.get("kwargs", {}))
#     # return __response.send_file(filename=comp.merge())
#     if cedula_type == "mercado":
#         query = {
#             "PRECIO_UNITARIO_APLICABLE",
#             "FOLIO",
#             "UBICACION_MANZANA",
#             "DIAS",
#             "CLASIFICACION_ECONOMICA",
#             "PRECIO_TOTAL_APLICABLE",
#             "NOMBRE_EDIFICIO",
#             "CALLE",
#             "TIPO_CONSUCCION",
#             "NUMERO_FRENTES",
#             "COORDENADAS_UTM_Y",
#             "CLASIFICACION_PERIFERICA",
#             "SUPERFICIE_CONSUCCION",
#             "ENE_CALLES",
#             "COLONIA",
#             "T_C",
#             "HOY",
#             "CADUCA_MESES",
#             "PRECIO_TOTAL_USD",
#             "EDAD",
#             "COORDENADAS_UTM_X",
#             "CALIDAD",
#             "FRENTE",
#             "UNIDADES_RENTABLES",
#             "EDO_CONSERVACION",
#             "DESCRIPCION_SERVICIOS",
#             "NIVELES",
#             "DESCRIPCION_ESPACIOS",
#             "SUPERFICIE_TERRENO",
#             "PROYECTO",
#         }
#     else:
#         query = {
#             "NOMBRE_VIALIDAD",
#             "COMERCIAL_FRENTE",
#             "COMPARABLE",
#             "SUPERFICIE",
#             "FRENTE_ML",
#             "NO_FRENTES",
#             "COORDENADA_X",
#             "ZONA_ECONOMICA",
#             "EDIFICIO_PREDIO_PROTOTIPO",
#             "PRECIO_USD",
#             "ENTRE_CALLES",
#             "INFRAESTRUCTURA",
#             "UBICACACION_MZA",
#             "ASENTAMIENTO",
#             "COORDENADA_Y",
#             "PERIFERIA",
#         }
#     comp.check(db=db)
#     return __response.success()


# @comparables.get("/cedulas/mercado")
# async def get_cedulas_mercado(
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get cedulas
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)

#     cedula = CedulaMercado(db)
#     if cedula.filter_group(usuario=str(user.id)) is None:
#         return __response.error(message="No se encontraron cedulas", status_code=404)
#     return __response.success(data=cedula.to_list())


# @comparables.get("/cedulas/comparable")
# async def get_cedulas_comparable(
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get cedulas
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)

#     cedula = CedulaComparables(db)

#     if cedula.all() is None:
#         return __response.error(message="No se encontraron cedulas", status_code=404)
#     return __response.success(data=cedula.to_list())


# @comparables.get("/cedula/mercado/{id}")
# async def get_cedula_mercado_by_id(
#     id: int,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get cedula by id
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)

#     cedula = CedulaMercado(db)

#     if cedula.filter(id=id, usuario=user.id) is None:
#         return __response.error(message="No se encontraron cedulas", status_code=404)
#     return __response.success(data=cedula.to_dict())


# @comparables.get("/cedula/comparable/{id}")
# async def get_cedula_comparable_by_id(
#     id: int,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Get cedula by id
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)
#     cedula = CedulaComparables(db)
#     if cedula.get(id) is None:
#         return __response.error(message="No se encontraron cedulas", status_code=404)
#     return __response.success(data=cedula.to_dict())


# @comparables.post("/cedula/mercado")
# async def create_cedula_mercado(
#     request: Request,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Create cedula
#     """
#     data = await request.json()
#     if isinstance(user, dict):
#         return __response.error(**user)
#     cedula = CedulaMercado(db)
#     if cedula.create(**data) is None:
#         return __response.error(message="No se pudo crear la cedula", status_code=404)
#     return __response.success(data=cedula.to_dict())


# @comparables.patch("/cedula/mercado/{id}")
# async def update_cedula_mercado(
#     id: int,
#     request: Request,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Update cedula
#     """
#     data = await request.json()
#     if isinstance(user, dict):
#         return __response.error(**user)
#     cedula = CedulaMercado(db)
#     if cedula.update(id, **data) is None:
#         return __response.error(
#             message="No se pudo actualizar la cedula", status_code=404
#         )
#     return __response.success(data=cedula.to_dict())


# @comparables.delete("/cedula/mercado/{id}")
# async def delete_cedula_mercado(
#     id: int,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     """
#     Delete cedula
#     """
#     if isinstance(user, dict):
#         return __response.error(**user)
#     cedula = CedulaMercado(db)
#     if cedula.delete(id) is None:
#         return __response.error(
#             message="No se pudo eliminar la cedula", status_code=404
#         )
#     return __response.success(data=cedula.to_dict())


# @comparables.post("/cedulas/mercado/{id}/preview/{comparable}")
# async def preview_cedulas(
#     request: Request,
#     id: int,
#     comparable: int,
#     db: Session = Depends(database.valuaciones),
#     user=Depends(required),
# ):
#     if isinstance(user, dict):
#         return __response.error(**user)
#     mercado = CedulaMercado(db)
#     cedula = CedulaComparables(db)
#     cat_com = ComparablesCatCom(db)
#     data = await request.json()
#     if mercado.filter(id=id, usuario=user.id) is None:
#         return __response.error(message="No se encontraron cedulas", status_code=404)
#     if cat_com.get(comparable) is None:
#         return __response.error(
#             message="No se encontraron comparables", status_code=404
#         )
#     if cedula.filter(id_cedula_mercado=id, id_comparable_catcom=comparable) is None:
#         if cedula.create(tipo=data.get("tipo", "TERRENO")) is None:
#             return __response.error(
#                 message="No se pudo crear la cedula", status_code=404
#             )
#     return
#     return
